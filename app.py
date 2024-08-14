from flask import Flask, request, jsonify
import pandas as pd
import openpyxl as xl
import os
from fpdf import FPDF
import matplotlib.pyplot as plt
import numpy as np
import json
import requests
from openpyxl.packaging import workbook

app = Flask(__name__)
arr_response=[]
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return 'No file part'

    file = request.files['file']
    if file.filename == '':
        return 'No selected file'

    file.save(file.filename)

    input_data = {
        'file_path': file.filename
    }
    file_path = os.path.abspath(input_data['file_path'])
    workbook = xl.load_workbook(file_path)
    num_sheets = len(workbook.sheetnames)
    result= {
        'num_of_sheets': num_sheets,
        'file_path': file_path
    }
    return f'Number of sheets: {result["num_of_sheets"]}, File path: {result["file_path"]}'


arr = []  # declare arr as a global variable


@app.route('/report', methods=['POST'])
def report():

    data = request.get_json()
    url = data.get('url')
    sheets = data.get('sheets')

    arr = []  # clearing the global variable before populating it

    for sheet in sheets:
        df = pd.read_excel(url, sheet_name=sheet['name'])
        column_sum = 0
        for column in sheet['columns']:
            if sheet['action'] == 'avg':
                column_sum += df[column].mean()
            if sheet['action'] == 'sum':
                column_sum += df[column].sum()
        arr.append(column_sum)
    global arr_response
    arr_response=arr
    generate_graphs(url, sheets)
    return jsonify(arr_response)


def convert_json_to_pdf(data):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    for idx,value in enumerate(data):
        pdf.cell(200, 10, f"Sheet Name: {idx+1} Result: {value} ", ln=True, align='C')

    pdf.output("complex_data.pdf")


def generate_graphs(excel_data, sheets):
   for sheet in sheets:
    df = pd.read_excel(excel_data, sheet_name=sheet['name'])

    plt.figure(figsize=(6, 4))
    for column in sheet['columns']:
        plt.plot(df[column], label=column)
    plt.xlabel('Rows')
    plt.ylabel('Values')
    plt.title(f'Graph for {sheet["name"]}')
    plt.legend()

    graph_filename = f"{sheet['name']}_graph.png"
    plt.savefig(graph_filename)

if __name__ == '__main__':
    app.run()
    convert_json_to_pdf(arr_response)



